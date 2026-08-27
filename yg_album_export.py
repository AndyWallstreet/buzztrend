# -*- coding: utf-8 -*-
"""YG 워크북에서 앨범 판매량 + 컨센서스 추이를 뽑아 data/yg/ 에 저장.

Usage:
    python yg_album_export.py
    (끝나면: git add data/yg && git commit && git push)

- Album 시트  : 연도 × 아티스트 실물 앨범 판매량(Circle Chart 집계)
- Consensus 시트 : 2026년 연간 매출/영업이익 컨센서스의 일별 추이 + 참여 증권사 수
투어 데이터는 yg_tour_export.py 가 따로 담당한다.
"""
import json
import shutil
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd

WB_PATH = Path(r"C:\Users\user99i1\LK자산운용\LK자산운용 - 문서\Companies\YG 엔터"
               r"\YG 엔터_Analysis template_2026 08_v02.xlsx")
DATA = Path(__file__).resolve().parent / "data" / "yg"


def open_workbook():
    """엑셀이 파일을 잠그고 있으면 임시 사본으로 읽는다."""
    try:
        return openpyxl.load_workbook(WB_PATH, read_only=True, data_only=True)
    except PermissionError:
        tmp = Path(tempfile.gettempdir()) / "yg_album_copy.xlsx"
        shutil.copy(WB_PATH, tmp)
        print(f"워크북이 잠겨 있어 사본으로 읽음: {tmp}")
        return openpyxl.load_workbook(tmp, read_only=True, data_only=True)


def export_album(wb) -> pd.DataFrame:
    """Album 시트 B4:K26 -> (year, artist, copies) 롱 포맷."""
    ws = wb["Album"]
    rows = list(ws.iter_rows(min_row=4, max_row=27, min_col=2, max_col=11,
                             values_only=True))
    artists = [a for a in rows[0][1:] if a]
    out = []
    for r in rows[1:]:
        yr_raw = r[0]
        if yr_raw is None:
            continue
        yr_s = str(yr_raw).strip()
        if not yr_s[:4].isdigit():
            continue
        for artist, v in zip(artists, r[1:1 + len(artists)]):
            if v is None or not isinstance(v, (int, float)) or v <= 0:
                continue
            out.append({"year": int(yr_s[:4]),
                        "is_est": yr_s.upper().endswith("E"),
                        # 컬럼 헤더의 '(copies)' 같은 꼬리표 제거
                        "artist": artist.split("(")[0].strip(),
                        "copies": float(v)})
    df = pd.DataFrame(out).sort_values(["year", "artist"]).reset_index(drop=True)
    df.to_csv(DATA / "album_sales.csv", index=False, encoding="utf-8")
    print(f"album_sales.csv — {df['year'].nunique()}개 연도, "
          f"{df['artist'].nunique()}팀, {len(df)}행")
    return df


def export_consensus(wb) -> pd.DataFrame:
    """Consensus 시트 -> 일별 2026E 매출/영업이익 컨센서스 (억원)."""
    ws = wb["Consensus"]
    base = ws.cell(12, 2).value or ""        # 예: 2026AS
    unit = str(ws.cell(11, 2).value or "")   # Local thou = 천원
    out = []
    for r in ws.iter_rows(min_row=15, max_row=ws.max_row, min_col=1, max_col=5,
                          values_only=True):
        d, rev, op, n_rev, n_op = r[0], r[1], r[2], r[3], r[4]
        if d is None or not isinstance(rev, (int, float)):
            continue
        out.append({"date": pd.Timestamp(d).date().isoformat(),
                    # 천원 단위 -> 억원
                    "rev_eok": rev / 1e5, "op_eok": (op or 0) / 1e5,
                    "n_rev": n_rev, "n_op": n_op})
    df = pd.DataFrame(out)
    df.to_csv(DATA / "consensus.csv", index=False, encoding="utf-8")
    meta = {"base": str(base), "unit_src": unit,
            "updated": str(ws.cell(1, 2).value or "")}
    (DATA / "consensus_meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"consensus.csv — {len(df)}일 ({df['date'].min()} ~ {df['date'].max()}), "
          f"기준 {base}")
    return df


def main():
    DATA.mkdir(parents=True, exist_ok=True)
    wb = open_workbook()
    export_album(wb)
    export_consensus(wb)
    print("이제: git add data/yg && git commit -m 'yg album/consensus' && git push")


if __name__ == "__main__":
    main()
