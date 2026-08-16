# -*- coding: utf-8 -*-
"""밸류 스크리너 데이터 갱신 — 엑셀이 업데이트되면 이 스크립트 한 번만 실행.

Usage:
    python screener_update.py [xlsx_path]
    (then: git add data/screener && git commit -m "screener data update" && git push)

xlsx_path를 주면 그 파일을 읽는다 (엑셀이 원본을 잠그고 있을 때 SaveCopyAs
사본을 넘기는 용도). 안 주면 아래 XLSX 기본 경로를 읽는다.

Reads the Capital IQ comparables workbook (Data sheet, CW:DJ block +
DY:EB forward-2026 block) and rewrites data/screener/screener_data.csv
+ meta.json.
"""
import json
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

XLSX = (r"C:\Users\user99i1\LK자산운용\LK자산운용 - 문서\Resources"
        r"\02_Industry Analysis\Comparables cap iq_value_AJ editing_v04.xlsx")
OUT = Path(__file__).resolve().parent / "data" / "screener"

COLS = ["ticker", "company", "roic_sg", "roe_sg", "ev_sales", "ev_ebit",
        "ev_ebitda", "per", "pbr", "sector", "industry_group", "industry",
        "primary_industry", "sic_industry"]
# 2026E 원본값 (Data DY:EB) — ev_sales 등 본 컬럼은 이제 '2026E 있으면 2026E,
# 없으면 LTM' 으로 엑셀에서 블렌딩된 값이고, 아래 raw 컬럼으로 어느 쪽인지 판별한다
FWD_RAW = {"ev_sales": 128, "ev_ebit": 129, "ev_ebitda": 130, "per": 131}
NUM_COLS = ["roic_sg", "roe_sg", "ev_sales", "ev_ebit", "ev_ebitda",
            "per", "pbr", "price", "mcap", "ev_fcf"]


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else XLSX
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb["Data"]

    rows = []
    for r in ws.iter_rows(min_row=9, max_col=133):
        if r[100].value is None:          # CW: ticker
            continue
        rec = {name: cell.value for name, cell in zip(COLS, r[100:114])}
        rec["price"] = r[9].value         # J: stock price
        rec["mcap"] = r[16].value         # Q: market cap (KRW mm)
        rec["ev_fcf"] = r[132].value      # EC: EV/FCF (LTM, evfcf_update.py가 채움)
        for name, idx in FWD_RAW.items():  # DY:EB — 2026E raw
            rec[f"_{name}_fwd"] = r[idx].value
        rows.append(rec)

    df = pd.DataFrame(rows)
    for c in NUM_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    # 멀티플별 기준 표시: 2026E 추정치가 있으면 "2026E", 없어서 LTM으로
    # 대체됐으면 "LTM" (PBR은 항상 현재 장부가라 src 컬럼 없음)
    for name in FWD_RAW:
        fwd = pd.to_numeric(df[f"_{name}_fwd"], errors="coerce")
        df[f"{name}_src"] = np.where(fwd.notna(), "2026E", "LTM")
        df.loc[df[name].isna(), f"{name}_src"] = ""
        df = df.drop(columns=[f"_{name}_fwd"])
    df = df[df["company"].apply(lambda x: isinstance(x, str) and len(x) > 1
                                and "Invalid" not in x)]
    df = df[df["sector"].apply(lambda x: isinstance(x, str) and len(x) > 1)]

    OUT.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUT / "screener_data.csv", index=False, encoding="utf-8-sig")
    n_fwd = int((df["ev_sales_src"] == "2026E").sum())
    meta = {"as_of": str(ws.cell(row=4, column=10).value)[:10],   # Data!J4
            "source": "Capital IQ comparables workbook",
            "basis": "FY2026E consensus, LTM fallback",
            "n_fwd": n_fwd,
            "n_companies": len(df)}
    (OUT / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=1),
                                   encoding="utf-8")
    print(f"OK — {len(rows)} raw rows, {meta['n_companies']} companies, "
          f"as of {meta['as_of']}")


if __name__ == "__main__":
    main()
