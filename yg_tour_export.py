# -*- coding: utf-8 -*-
"""YG 워크북 → data/yg/ 투어 데이터 내보내기.

YG 엔터_Analysis template 워크북의 'Tour detail' 시트에서 BIGBANG(2026~)과
BABYMONSTER(2026~) 스탑을 뽑아 CSV 로 저장한다. 읽기 전용 — 워크북은 절대
저장하지 않는다 (FactSet 수식 보호). 사용자가 Excel 로 열어 잠겨 있으면
PowerShell Copy-Item 으로 사본을 떠서 읽는다 (python open 은 거부되지만
OS 복사는 된다 — 검증됨).

산출물 (data/yg/):
  bigbang_tour.csv / babymonster_tour.csv — 스탑별 일정·좌석·가격·매출 추정
  assumptions.json — 가격·taking rate·FX 등 모델 가정

워크북에서 투어 일정이 바뀌면 이 스크립트만 다시 돌리면 된다:
  python yg_tour_export.py
"""
import json
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

WB_PATH = Path(r"C:\Users\user99i1\LK자산운용\LK자산운용 - 문서\Companies\YG 엔터"
               r"\YG 엔터_Analysis template_2026 08_v02.xlsx")
DATA = Path(__file__).resolve().parent / "data" / "yg"

# Assumptions 시트와 동일 (바뀌면 여기도 갱신)
PRICE = {"KR": 100, "JP": 125, "NA/EU": 200, "Others": 80}
TAKING = {"KR": 0.40, "JP": 0.30, "NA/EU": 0.35, "Others": 0.35}
FX = 1500
FILL = 0.95

TARGETS = {  # group -> 시작 연도 (이후 전부)
    "BIGBANG": 2026,
    "BABYMONSTER": 2026,
}


def open_workbook():
    try:
        return openpyxl.load_workbook(WB_PATH, read_only=True, data_only=True)
    except PermissionError:
        tmp = Path(tempfile.gettempdir()) / "yg_export_copy.xlsx"
        subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             f'Copy-Item "{WB_PATH}" "{tmp}" -Force'],
            check=True)
        print(f"워크북이 잠겨 있어 사본으로 읽음: {tmp}")
        return openpyxl.load_workbook(tmp, read_only=True, data_only=True)


def main():
    DATA.mkdir(parents=True, exist_ok=True)
    wb = open_workbook()
    ws = wb["Tour detail"]

    out = {g: [] for g in TARGETS}
    for r in ws.iter_rows(min_row=4, values_only=True):
        group, tour, year = r[0], r[1], r[2]
        if group not in TARGETS or not year or year < TARGETS[group]:
            continue
        date, city, country, region, venue = r[6], r[7], r[8], r[9], r[10]
        cap, shows = r[11], r[12]
        if not date or not cap or not shows:
            continue
        seats = int(cap) * int(shows)
        price = PRICE[region]
        taking = TAKING[region]
        att = round(seats * FILL)
        gross_m = att * price / 1e6
        yg_m = gross_m * taking
        out[group].append({
            "group": group, "tour": tour,
            "date": date.date().isoformat(), "month": r[4], "quarter": r[5],
            "city": city, "country": country, "region": region, "venue": venue,
            "capacity": int(cap), "shows": int(shows), "seats": seats,
            "price_usd": price, "taking": taking,
            "att_est": att,
            "gross_usd_m": round(gross_m, 4),
            "yg_rev_usd_m": round(yg_m, 4),
            "yg_rev_krw_mn": round(yg_m * FX, 1),
            "ref": r[25] or "",
        })

    cols = ["group", "tour", "date", "month", "quarter", "city", "country",
            "region", "venue", "capacity", "shows", "seats", "price_usd",
            "taking", "att_est", "gross_usd_m", "yg_rev_usd_m",
            "yg_rev_krw_mn", "ref"]
    names = {"BIGBANG": "bigbang_tour.csv", "BABYMONSTER": "babymonster_tour.csv"}
    for g, rows in out.items():
        rows.sort(key=lambda x: x["date"])
        p = DATA / names[g]
        with p.open("w", encoding="utf-8-sig", newline="") as f:
            f.write(",".join(cols) + "\n")
            for row in rows:
                f.write(",".join(f'"{row[c]}"' if isinstance(row[c], str) and "," in row[c]
                                 else str(row[c]) for c in cols) + "\n")
        tot = sum(x["yg_rev_krw_mn"] for x in rows)
        print(f"{p.name}: {len(rows)}개 스탑 · YG 매출 추정 합계 ₩{tot:,.0f}mn")

    (DATA / "assumptions.json").write_text(json.dumps({
        "price_usd": PRICE, "taking": TAKING, "fx": FX, "fill": FILL,
        "source": WB_PATH.name,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print("assumptions.json written")


if __name__ == "__main__":
    main()
